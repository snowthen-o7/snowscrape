"""
Format converter for exporting scraping results to multiple formats.

Supports CSV, SQL, and XLSX. CSV and SQL use only the standard library; XLSX
uses openpyxl (a pure-Python dependency that IS bundled in the Lambda). Parquet
requires pyarrow, which is intentionally NOT bundled (size / cold-start cost), so
a Parquet request raises a clear, handled error instead of crashing.

History (issue #32): this module previously imported pandas and openpyxl at the
module top. pandas is deliberately not bundled in the Lambda (see
utils.parse_links_from_file and backend/pyproject.toml), so the import always
raised ModuleNotFoundError in production, and because the download handler
imported the converter outside its conversion try/except, EVERY non-JSON export
(CSV/XLSX/Parquet/SQL) crashed with an unhandled 500. The converters are now
pandas-free and resolve the S3 client lazily (the old module-level
boto3.client('s3') pinned credentials at import time, which also breaks under
moto in tests).
"""

import csv
import json
from io import BytesIO, StringIO
from typing import Dict, List, Optional

from connection_pool import get_s3_client
from logger import get_logger

logger = get_logger(__name__)


def _scalarize(value) -> str:
    """Render a single cell value as a string for text/tabular formats.

    - None -> '' (SQL handles None as NULL before calling this).
    - list -> pipe-joined (matches the flatten convention used elsewhere).
    - dict / other nested structures -> compact, deterministic JSON.
    - scalars -> str().
    """
    if value is None:
        return ''
    if isinstance(value, list):
        return '|'.join(str(v) for v in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


# Leading characters that Excel / Google Sheets / LibreOffice may interpret as
# the start of a formula. Scraped cell values are attacker-influenced, so a value
# beginning with one of these is prefixed with a single quote to force literal
# text rendering (spreadsheet/CSV formula injection, CWE-1236; issue #34).
_FORMULA_TRIGGERS = ('=', '+', '-', '@', '\t', '\r')


def _sanitize_spreadsheet_text(text: str) -> str:
    """Neutralize spreadsheet formula injection in an already-stringified cell.

    Prefixes a single quote when the value begins with a formula trigger so the
    spreadsheet renders it as text instead of evaluating it. Applied only to the
    CSV and XLSX (string) write paths; SQL and Parquet are intentionally
    unaffected (SQL emits single-quote-escaped string literals, Parquet is a
    binary data format, neither is formula-evaluated by a spreadsheet app).
    """
    if text and text[0] in _FORMULA_TRIGGERS:
        return "'" + text
    return text


def _sql_ident(name) -> str:
    """Backtick-quote a SQL identifier, escaping embedded backticks.

    Column names come from arbitrary scraped JSON keys, so a key containing a
    backtick must not be allowed to break out of the identifier quoting.
    """
    return "`" + str(name).replace("`", "``") + "`"


def _xlsx_cell(value):
    """Coerce a value into something openpyxl can write to a cell.

    Numbers/bools stay native so they remain numeric in Excel; None becomes ''
    and nested values are scalarized to a stable string. String cells (including
    scalarized nested values) are run through the formula-injection guard so a
    scraped string like "=1+1" is written as literal text, not a formula
    (openpyxl would otherwise store a leading-'=' string as a formula cell).
    """
    if value is None:
        return ''
    if isinstance(value, (list, dict)):
        return _sanitize_spreadsheet_text(_scalarize(value))
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        return _sanitize_spreadsheet_text(value)
    return _sanitize_spreadsheet_text(str(value))


class FormatConverter:
    """Converts job results from JSON to various export formats."""

    def __init__(self, job_id: str, results_data: List[Dict], s3_bucket: str):
        """
        Initialize the format converter.

        Args:
            job_id: The job identifier
            results_data: List of result dictionaries from JSON
            s3_bucket: S3 bucket name for storing converted files
        """
        self.job_id = job_id
        self.results_data = results_data or []
        self.s3_bucket = s3_bucket
        # Populated by prepare_dataframe(): the tabular rows + their ordered columns.
        self.rows: Optional[List[Dict]] = None
        self.columns: List[str] = []

    def prepare_dataframe(self, flatten: bool = False) -> List[Dict]:
        """
        Normalize JSON results into rows + an ordered column list.

        The method name is retained for call-site compatibility with the
        download handler (it no longer returns a Pandas DataFrame).

        Args:
            flatten: If True, flatten nested query results into columns
                     (url/status/http_code + each data key); otherwise keep the
                     raw row dicts.

        Returns:
            The prepared list of row dicts.
        """
        rows: List[Dict] = []

        if flatten:
            for row in self.results_data:
                if not isinstance(row, dict):
                    continue
                flat_row = {
                    'url': row.get('url', ''),
                    'status': row.get('status', ''),
                    'http_code': row.get('http_code', ''),
                }
                data = row.get('data', {})
                if isinstance(data, dict):
                    for key, value in data.items():
                        flat_row[key] = value
                rows.append(flat_row)
        else:
            rows = [dict(row) for row in self.results_data if isinstance(row, dict)]

        # Column order = first-seen union across all rows (matches the old
        # pd.DataFrame(list_of_dicts) column ordering, which exposed the union
        # of keys in first-appearance order).
        columns: List[str] = []
        seen = set()
        for row in rows:
            for key in row.keys():
                if key not in seen:
                    seen.add(key)
                    columns.append(key)

        self.rows = rows
        self.columns = columns

        logger.info("Rows prepared", job_id=self.job_id, rows=len(rows),
                    columns=len(columns), flatten=flatten)
        return self.rows

    def _ensure_prepared(self) -> None:
        """Lazily prepare rows (flattened) if the handler did not call prepare first."""
        if self.rows is None:
            self.prepare_dataframe(flatten=True)

    def convert_to_csv(self, s3_key: str) -> str:
        """
        Convert to CSV (standard-library csv writer) and upload to S3.

        Args:
            s3_key: S3 key for the CSV file

        Returns:
            S3 key of uploaded file
        """
        self._ensure_prepared()

        buffer = StringIO()
        # lineterminator='\n' keeps output deterministic (csv defaults to '\r\n').
        writer = csv.DictWriter(
            buffer, fieldnames=self.columns, extrasaction='ignore', lineterminator='\n'
        )
        # Header cells come from scraped/AI-derived JSON keys too, so they get
        # the same formula-injection guard as data cells (not writer.writeheader).
        writer.writerow({col: _sanitize_spreadsheet_text(col) for col in self.columns})
        for row in self.rows:
            writer.writerow(
                {col: _sanitize_spreadsheet_text(_scalarize(row.get(col)))
                 for col in self.columns}
            )

        body = buffer.getvalue().encode('utf-8')

        get_s3_client().put_object(
            Bucket=self.s3_bucket,
            Key=s3_key,
            Body=body,
            ContentType='text/csv',
            Metadata={
                'job-id': self.job_id,
                'format': 'csv',
                'row-count': str(len(self.rows)),
            },
        )

        logger.info("CSV file uploaded to S3", job_id=self.job_id, s3_key=s3_key,
                    rows=len(self.rows))
        return s3_key

    def convert_to_xlsx(self, s3_key: str) -> str:
        """
        Convert to Excel (XLSX) with formatting via openpyxl and upload to S3.

        Args:
            s3_key: S3 key for the XLSX file

        Returns:
            S3 key of uploaded file
        """
        self._ensure_prepared()

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as exc:  # pragma: no cover - openpyxl is bundled
            raise RuntimeError(
                "XLSX export unavailable: openpyxl is not installed in this environment"
            ) from exc

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Job {str(self.job_id)[:8]}"

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, column in enumerate(self.columns, 1):
            # Header text is scraped/AI-derived too -> same formula guard.
            cell = ws.cell(row=1, column=col_idx,
                           value=_sanitize_spreadsheet_text(str(column)))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for row_idx, row in enumerate(self.rows, 2):
            for col_idx, column in enumerate(self.columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=_xlsx_cell(row.get(column)))

        # Auto-size columns (max width 50)
        for col_idx, column in enumerate(self.columns, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(column))
            for row in self.rows:
                value = row.get(column)
                if value is not None and value != '':
                    max_length = max(max_length, len(_scalarize(value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Auto-filter + frozen header (only meaningful when there are columns)
        if self.columns:
            ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        get_s3_client().put_object(
            Bucket=self.s3_bucket,
            Key=s3_key,
            Body=excel_buffer.getvalue(),
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            Metadata={
                'job-id': self.job_id,
                'format': 'xlsx',
                'row-count': str(len(self.rows)),
            },
        )

        logger.info("XLSX file uploaded to S3", job_id=self.job_id, s3_key=s3_key,
                    rows=len(self.rows))
        return s3_key

    def convert_to_parquet(self, s3_key: str) -> str:
        """
        Convert to Parquet with Snappy compression and upload to S3.

        Requires pyarrow, which is NOT bundled in the Lambda (size / cold-start
        cost). When pyarrow is unavailable this raises a clear RuntimeError that
        the download handler turns into a clean 500 instead of an unhandled
        ModuleNotFoundError.

        Args:
            s3_key: S3 key for the Parquet file

        Returns:
            S3 key of uploaded file
        """
        self._ensure_prepared()

        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except ImportError as exc:
            raise RuntimeError(
                "Parquet export unavailable: pyarrow is not installed in this "
                "environment (not bundled in the Lambda due to size). Use CSV, "
                "XLSX, or SQL instead."
            ) from exc

        # All columns are emitted as strings to keep the schema stable across
        # heterogeneous JSON rows.
        table = pa.table(
            {col: [_scalarize(row.get(col)) for row in self.rows] for col in self.columns}
        )
        parquet_buffer = BytesIO()
        pq.write_table(table, parquet_buffer, compression='snappy')
        parquet_buffer.seek(0)

        get_s3_client().put_object(
            Bucket=self.s3_bucket,
            Key=s3_key,
            Body=parquet_buffer.getvalue(),
            ContentType='application/octet-stream',
            Metadata={
                'job-id': self.job_id,
                'format': 'parquet',
                'row-count': str(len(self.rows)),
                'compression': 'snappy',
            },
        )

        logger.info("Parquet file uploaded to S3", job_id=self.job_id, s3_key=s3_key,
                    rows=len(self.rows))
        return s3_key

    def convert_to_sql(self, s3_key: str, table_name: str = 'scraped_data') -> str:
        """
        Generate SQL INSERT statements (standard-library only) and upload to S3.

        Args:
            s3_key: S3 key for the SQL file
            table_name: Name of the SQL table

        Returns:
            S3 key of uploaded file
        """
        self._ensure_prepared()

        if not self.columns:
            # No columns => no data (empty or all-non-dict results). Emit a valid
            # SQL comment instead of `CREATE TABLE ... ()`, which is invalid DDL.
            sql_content = (
                f"-- No data to export for table {_sql_ident(table_name)} "
                f"(0 columns, {len(self.rows)} rows)\n"
            )
        else:
            quoted_table = _sql_ident(table_name)
            sql_statements = []

            # CREATE TABLE with TEXT columns (identifiers backtick-escaped)
            col_defs = ', '.join(f"{_sql_ident(col)} TEXT" for col in self.columns)
            sql_statements.append(
                f"CREATE TABLE IF NOT EXISTS {quoted_table} ({col_defs});\n\n"
            )

            # INSERT statements in batches of 100 rows
            if self.rows:
                column_list = ', '.join(_sql_ident(col) for col in self.columns)
                batch_size = 100
                for i in range(0, len(self.rows), batch_size):
                    batch = self.rows[i:i + batch_size]
                    values = []
                    for row in batch:
                        escaped = []
                        for col in self.columns:
                            v = row.get(col)
                            if v is None:
                                escaped.append('NULL')
                            else:
                                # Escape single quotes by doubling them
                                escaped.append("'" + _scalarize(v).replace("'", "''") + "'")
                        values.append(f"({', '.join(escaped)})")

                    insert_stmt = f"INSERT INTO {quoted_table} ({column_list}) VALUES\n"
                    insert_stmt += ',\n'.join(values) + ";\n\n"
                    sql_statements.append(insert_stmt)

            sql_content = ''.join(sql_statements)

        get_s3_client().put_object(
            Bucket=self.s3_bucket,
            Key=s3_key,
            Body=sql_content.encode('utf-8'),
            ContentType='application/sql',
            Metadata={
                'job-id': self.job_id,
                'format': 'sql',
                'table-name': table_name,
                'row-count': str(len(self.rows)),
            },
        )

        logger.info("SQL file uploaded to S3", job_id=self.job_id, s3_key=s3_key,
                    table_name=table_name, rows=len(self.rows))
        return s3_key
